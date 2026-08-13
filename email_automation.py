"""
Outlook Email -> Excel Exporter (Daily Folder + One Row Per Conversation)
--------------------------------
Does exactly two things:
    1. Reads every email received on the target date (see DAYS_OFFSET) in
       the given Outlook folder (default: Inbox), moves each one into a
       folder under Inbox named with that date (e.g. '2026-07-14'), and
       writes an Excel export. If a conversation already has an EARLIER
       date folder (e.g. it started on 14-July and the customer replied
       again on 15-July), the new message is merged into that ORIGINAL
       folder instead of creating an entry in today's folder.
    2. Excel has ONE ROW PER CONVERSATION, not per email. If a
       conversation had several back-and-forth exchanges that day, that
       one row gets repeating "Receiving Time N" / "Response Time N"
       column pairs (N = 1, 2, 3...) instead of separate rows. If a
       conversation only had a single message with no reply yet, only
       pair 1 is filled and the rest of that row's pair columns show "-".

Column basis for timing:
    - "Receiving Time N" = when the Nth customer message arrived
      (msg.ReceivedTime)
    - "Response Time N"  = when the FIRST reply after that message was
      actually sent (from Sent Items, msg.SentOn) - "-" if no reply yet

Each conversation also gets a "Status" field:
    "-"       -> default, we don't have a verified customer number yet
    "Open"    -> we have their number and the complaint is still active
    "Resolve" -> the complaint has been marked resolved
Status gates both HOD escalation and the follow-up nudge - only "Open"
conversations get escalated or nudged; "Resolve" ones are left alone.

Requirements (run once):
    pip install pywin32 openpyxl

Requirements:
    - Outlook desktop app must be installed and configured with your account
      (Windows only - this uses COM automation, so it will NOT work on Mac/Linux)
    - Outlook can be open or closed; the script will launch/attach to it automatically
    - The target date's folder is created automatically under Inbox if it
      doesn't already exist - no manual setup needed.

Usage:
    python outlook_to_excel.py

CHANGELOG (fixes applied on top of the original version):
    1. "safety" category keywords no longer overlap with URGENT_KEYWORDS,
       so urgent-but-unrelated emails ("need my invoice ASAP") are no
       longer misclassified as safety complaints.
    2. Draft-reply logic is now mutually exclusive and ordered
       resolve -> missing number -> normal reply, instead of three
       independent checks that could silently fight each other.
    3. HOD escalation emails only mark a conversation as "escalated" if
       mail.Send() actually succeeded - a failed send no longer
       permanently suppresses a real escalation.
    4. The "long conversation" escalation trigger now fires on threads
       with a genuine unresolved backlog, not on healthy, fully-answered
       long threads.
    5. Auto-resolve detection ("thanks", "done", etc.) is skipped for
       safety-category conversations, so a message like "thanks for
       nothing, brakes still not fixed" can't auto-close a safety case.
    6. Sent Items and historical date folders are each scanned ONCE per
       run (cached into conv_id -> messages maps) instead of being
       rescanned in full for every single conversation - this removes
       an O(conversations x mailbox size) performance problem.
    7. refresh_outlook() only runs under `if __name__ == "__main__":`,
       so importing this module no longer triggers a live Outlook
       send/receive as a side effect.
    8. HOD email / escalation threshold / target folder are now read
       from environment variables (with sensible defaults), instead of
       being hardcoded values buried in the middle of the file.
    9. Added an explicit per-conversation Status field ("-" / "Open" /
       "Resolve") that gates both HOD escalation and follow-up nudges,
       instead of scattering "not is_resolved" checks everywhere.
    10. Once a conversation has ever been marked "Resolve", it stays
        closed for drafting purposes: no further auto-draft replies
        (resolve thank-you, missing-number ask, or category reply) are
        generated for ANY later message in that conversation, the same
        way a closed ticket in a real helpdesk doesn't keep auto-replying
        just because the customer's mailbox thread got another message.
        The resolved state is persisted on the Outlook item itself via a
        Category tag, so it survives across runs/days.
"""

import os
import re
import time
from datetime import datetime, timedelta

import openpyxl
import win32com.client
from openpyxl.utils import get_column_letter


# --------------------------------------------------------------------------
# CONFIG - override any of these via environment variables without touching
# code, e.g.  set HOD_ESCALATION_EMAIL=hod@yourcompany.com
# --------------------------------------------------------------------------

#------------------------Internal MAILS FEATURE ---------------------------------------------------
#Send this to true and set the emails to get the Internal emails variant

OUTLOOK_FOLDER_NAME = os.environ.get("OUTLOOK_FOLDER", "Inbox")                 #Folder from which we read all emails and in which we create new folders

DATE_FOLDER_PATTERN = re.compile(r"^\d{4}-\d{2}-\d{2}$")

#1st variable
#ENABLE_INTERNAL_MAILS = os.environ.get("ENABLE_INTERNAL_MAILS", "False").lower() == "true"  # Set INTERNAL email (True or false)

#1st var NO use
#Special_Email_NDays = int(os.environ.get("Special_Email_NDays", 8))              #INTERNAL Email threshold

#2nd var
#2nd var - whole domains only, no "@", e.g. "abc.com.pk"
INTERNAL_MAIL_DOMAINS = [
    "abc.com.pk",
    "microsoft.com"
    # add as many domains as needed
]

#3rd VAR
ENABLE_FOLDER_CLEANUP = os.environ.get("ENABLE_FOLDER_CLEANUP", "true").lower() == "true"

#4th VAR
DELETE_OLDER_THAN_DAYS = int(os.environ.get("DELETE_OLDER_THAN_DAYS", "15"))      #Nth day folder you want to delete

#5TH VAR
NO_OF_DAYS = int(os.environ.get("NO_OF_DAYS", 8))                               #NO of days excel sheet & Folders you want.

#6TH VAR
HOD_EMAIL = os.environ.get("HOD_EMAIL", "k233001@nu.edu.pk")   #On escalation send an email to this

#7TH VAR    
ESCALATION_THRESHOLD = int(os.environ.get("ESCALATION_THRESHOLD", "5"))           #Threshold after which send email to HOD

#8TH VAR
FOLLOWUP_DAYS_THRESHOLD = int(os.environ.get("FOLLOWUP_DAYS_THRESHOLD", "3"))  # FOLLOW UP THRESHOLD: NO OF DAYS 

#9TH VAR
IS_CUSTOMER_CARE = os.environ.get("IS_CUSTOMER_CARE", "true").lower() == "true"

#10TH VAR
DAYS_OFFSET = int(os.environ.get("DAYS_OFFSET", "0"))  # 0 for today, 1 for yesterday, etc.

#11Th VAR ---------------- PROMOTION ---------------------- if promotion -> then move to Promotions
promotions = ["free charger", "free bag", "promotion", "promotions"]

#12TH VAR --------------- MULTI-DATE FOLDER SUPPORT ----------------------
# When true (default), a run scans the WHOLE Inbox, discovers every date
# that has mail sitting in it, and creates/uses a folder per date instead
# of only handling the single DAYS_OFFSET date. Set to "false" to restore
# the original single-date-only behaviour.
PROCESS_ALL_INBOX_DATES = os.environ.get("PROCESS_ALL_INBOX_DATES", "true").lower() == "true"

#13TH VAR --------------- HOD OVERDUE ALERT (WORKING DAYS) ----------------
# If a customer message sits unanswered for this many WORKING days
# (Mon-Fri only, weekends don't count), HOD gets an alert email.
HOD_POPUP_BUSINESS_DAYS = int(os.environ.get("HOD_POPUP_BUSINESS_DAYS", "2"))



#SET_N_DAYS = int(os.environ.get("SET_N_DAYS", "8"))   #By default I set it to 8  #I set everything to this #Not use anywhere as everything has its own value

PROJECT_DIRECTORY = r"C:\Users\Abdullahh\OneDrive\Desktop\Anaconda_Project\firstproject\Outlook_files" #Diectory for Excel Files

OL_FOLDER_INBOX = 6

outlook = win32com.client.Dispatch("Outlook.Application")

PERSONAL_EMAIL = outlook.Session.Accounts.Item(1).SmtpAddress







def get_target_date():
    return (datetime.now() - timedelta(days=DAYS_OFFSET))

def add_business_days(start_dt, n):
    """Adds n WEEKDAY days to start_dt, keeping the same time-of-day,
    skipping Sat/Sun entirely. e.g. Monday 10:00 AM + 2 -> Wednesday
    10:00 AM. Friday 10:00 AM + 2 -> Tuesday 10:00 AM (Sat/Sun skipped)."""
    current = start_dt
    added = 0
    while added < n:
        current += timedelta(days=1)
        if current.weekday() < 5:  # Mon=0 ... Fri=4
            added += 1
    return current


# Category tag used to persist "this conversation has been resolved" on
# the actual Outlook item (Categories field), so the closed state
# survives across runs/days without needing an external database.
RESOLVED_CATEGORY_TAG = "ResolvedConversation"

HOD_POPUP_TAG_PREFIX = "HodPopupSent|"


def build_hod_popup_sent_map(outlook_namespace):
    """conv_id -> set of customer-message-timestamp strings that already
    triggered an overdue alert, so the same unanswered message doesn't
    re-alert HOD every single run."""
    sent_folder = outlook_namespace.GetDefaultFolder(5)  # Sent Items
    popup_map = {}
    for item in sent_folder.Items:
        try:
            if check_outlook_item(item.Class):
                continue
            categories = item.Categories or ""
            if categories.startswith(HOD_POPUP_TAG_PREFIX):
                parts = categories.split("|")
                if len(parts) == 3:
                    _, conv_id, ts = parts
                    popup_map.setdefault(conv_id, set()).add(ts)
        except Exception:   
            continue
    return popup_map

# --- Rule-based complaint classification (tuned for a motor company) ---
# NOTE: "safety" intentionally does NOT include generic urgency words
# ("urgent", "asap", "immediately", ...) - those live only in
# URGENT_KEYWORDS below. Mixing them in here previously caused any email
# containing "ASAP" to be misclassified as a safety complaint regardless
# of content.
CATEGORY_KEYWORDS = {
    "safety": [
        "accident", "brake failure", "brake fail", "airbag", "steering fail",
        "smoke", "fire", "crash", "unsafe", "collision", "explode",
        "explosion", "burst tire", "burst tyre", "seatbelt", "abs fail",
        "loss of control", "life threatening",
    ],
    "mechanical": [
        "engine failure", "engine stall", "won't start", "not starting",
        "transmission", "gearbox", "clutch", "overheating", "oil leak",
        "coolant leak", "battery dead", "suspension issue", "electrical fault",
        "strange noise", "grinding noise", "stalling",
    ],
    "warranty": [
        "warranty", "defect", "manufacturing issue", "replace part",
        "faulty part", "recall", "factory fault",
    ],
    "service_delay": [
        "delay", "still not fixed", "waiting", "late", "appointment",
        "not repaired", "no response", "follow up",
    ],
}

# Urgency is scored independently of category - it only affects PRIORITY,
# never which category an email is filed under.
URGENT_KEYWORDS = [
    "urgent", "critical", "asap", "immediately", "emergency",
    "severe", "high priority", "escalate", "right away", "life threatening",
]

REPLY_TEMPLATES = {
    "safety": (
        "Dear {name},\n\n"
        "Thank you for reporting this safety concern. This has been escalated to "
        "our Safety & Quality team, and a specialist will contact you within 24 hours."
    ),
    "mechanical": (
        "Dear {name},\n\n"
        "Thank you for reporting this mechanical issue. Our service team has "
        "logged your case and will contact you to schedule an inspection."
    ),
    "warranty": (
        "Dear {name},\n\n"
        "Thank you for contacting us about your warranty claim. Our service team "
        "will review your case and respond within 2 business days."
    ),
    "service_delay": (
        "Dear {name},\n\n"
        "We're sorry to hear about the delay you've experienced. Our service "
        "center manager will reach out to you shortly."
    ),
    "general": (
        "Dear {name},\n\n"
        "Thank you for reaching out. We've logged your message and a representative "
        "will respond within 1-2 business days."
    ),
    "Missing_Num": (
        "Dear {name}, \n\n"
        "Thank you for your email to Master Changan Motors Limited."
        "With reference to your concern, we kindly request you to share your vehicle details, including the chassis number, registration number, dealership name, current mileage, and an active contact number, so that we may proceed further accordingly."
        "For any additional assistance, please feel free to contact us at 021-111-116-265 between 7:45 AM and 4:15 PM.\n"
        "Kindly provide a valid Pakistani contact number (e.g., 03XXXXXXXXX or +923XXXXXXXXX).\n"
    ),
    "Complain_Resolve": (
        "Dear {name}, \n\n"
        "We are pleased to inform you that your reported concern has been resolved."
        "Thank you for bringing this matter to our attention and for your patience throughout the resolution process. We hope the solution provided has addressed your concern satisfactorily."
        "If you experience any further issues or require additional assistance, please don't hesitate to contact us. We will be happy to assist you."
        "Thank you for choosing Master Changan Motors Limited.\n"
        "Kind regards,\n"
        "Customer Experience Team\n"
        "Master Changan Motors Limited\n"
    ),
    "Followup_Nudge": (
        "Dear {name},\n\n"
        "We wanted to follow up on your recent inquiry, as we haven't heard back from "
        "you in a while and your concern is still open on our end. If you still need "
        "assistance, please reply to this email at your earliest convenience so we "
        "can continue helping you.\n\n"
        "If your issue has already been resolved, please let us know so we can close "
        "this out.\n\n"
        "Regards,\nCustomer Experience Team\nMaster Changan Motors Limited"
    )
}

# Split into STRONG phrases (unambiguous - nobody says "issue resolved"
# unless it is) vs WEAK words (common courtesy language that shows up
# constantly even when the customer has an open ask, e.g. "thanks" at
# the top of "thanks for your response, but I want to ask...").
STRONG_RESOLVED_PHRASES = [
    "issue resolved", "issue is resolved", "issue has been resolved",
    "problem solved", "problem is solved",
    "you may close", "close the complaint", "close this complaint",
    "resolved my issue", "resolved my problem", "resolved my complaint",
]
WEAK_RESOLVED_WORDS = [
    "resolved", "fixed", "working now", "thank you", "thanks",
    "okay now", "ok now", "satisfied", "done", "completed",
]

# A resolved-sounding word right next to a contrast conjunction almost
# always means the customer is pivoting to something else -
# "thanks for your help, BUT I still have a question about..."
CONTRAST_WORDS = ["but ", " but,", "however", "although", "even though", "except that", "on the other hand"]

# Markers where Outlook's quoted-reply history typically begins in plain
# text bodies - everything from the first matching marker onward is
# treated as quoted history, not the customer's new text, so resolve
# detection never gets poisoned by our OWN template text sitting in the
# quoted thread below a customer's new reply.
QUOTE_MARKERS = ["-----Original Message-----", "\nFrom:", "\r\nFrom:", "\nOn ", "\r\nOn "]


def refresh_outlook():
    outlook = win32com.client.Dispatch("Outlook.Application")
    namespace = outlook.GetNamespace("MAPI")
    print("Refreshing Outlook...")
    print(f"Using Email : {PERSONAL_EMAIL}")  # Print the email address of the specific email
    print(f"HOD Email   : {HOD_EMAIL}")       #Print the HOD Email

    namespace.SendAndReceive(True)
    time.sleep(2)
    print("Refresh Complete.")


import win32com.client


def move_all_spam_to_inbox():
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

    # Default folders
    inbox = outlook.GetDefaultFolder(6)   # olFolderInbox
    junk = outlook.GetDefaultFolder(23)   # olFolderJunk

    items = junk.Items

    # Iterate backwards because the collection changes as items are moved
    for i in range(items.Count, 0, -1):
        try:
            items.Item(i).Move(inbox)
        except Exception as e:
            print(f"Failed to move email: {e}")

    print("All Spam/Junk emails moved to Inbox.")

def check_promotion(subject, body):
    text = (subject or "").lower() + (body or "").lower()
    return any(promo.lower() in text for promo in promotions)
# ----------------------------------------------------------------------------------


# Only read Primary emails
def is_primary(msg):
    """Filters out obvious bulk/promotional mail via the list-unsubscribe header."""
    try:
        header = msg.PropertyAccessor.GetProperty(
            "http://schemas.microsoft.com/mapi/proptag/0x007D001E"
        )
        if "list-unsubscribe" in header.lower():
            return False
    except Exception:
        pass
    return True

def is_internal_domain(sender_email, allowed_domains_lower):
    """True if sender_email's domain is in INTERNAL_MAIL_DOMAINS
    (e.g. 'abc.com.pk' matches 'anyone@abc.com.pk')."""
    sender_email = (sender_email or "").lower().strip()
    if "@" not in sender_email:
        return False
    domain = sender_email.rsplit("@", 1)[-1]
    return domain in allowed_domains_lower

# getting the sender email
def get_sender_email(msg):
    """Handles both Exchange and SMTP sender address formats."""
    try:
        if msg.SenderEmailType == "EX":
            return msg.Sender.GetExchangeUser().PrimarySmtpAddress
        return msg.SenderEmailAddress
    except Exception:
        return msg.SenderEmailAddress or ""


# Get Message id
def get_message_id(msg):
    try:
        return msg.PropertyAccessor.GetProperty(
            "http://schemas.microsoft.com/mapi/proptag/0x1035001F"
        )
    except Exception:
        return ""


# Get folder from application, Default = Inbox.
def get_outlook_folder(folder_name="Inbox"):
    """Connects to Outlook and returns the requested folder (default Inbox)."""
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = outlook.GetDefaultFolder(OL_FOLDER_INBOX)  # 6 = olFolderInbox

    if folder_name.lower() == "inbox":
        return inbox

    for folder in inbox.Folders:  # find in Inbox
        if folder.Name.lower() == folder_name.lower():
            return folder

    print(f"Folder '{folder_name}' not found under Inbox. Using Inbox instead.")
    return inbox


#Check whether if outlook item is email ?
#return true if not an email
#if class == 43 --> then an email else not a normal email
def check_outlook_item(Msg_class):
    if Msg_class != 43: #not an email  (43 is an email only so we can read it & ignore all email where class != 43)
        return True
    return False



# Create Subfolder if not found in parent folder, eg for today, we sent (inbox, today_date)
def get_or_create_subfolder(parent_folder, name):
    """Returns the subfolder with this name under parent_folder, creating it if missing."""
    for f in parent_folder.Folders:
        if f.Name == name:
            return f
    print(f"Creating folder '{name}' under '{parent_folder.Name}'...")
    return parent_folder.Folders.Add(name)


# ---------- FOR PROMOTIONS RELATED EMAILS
def setup_promotions_folder(inbox_folder):
    """Creates (or reuses) a 'Promotions' folder directly under Inbox."""
    return get_or_create_subfolder(inbox_folder, "Promotions")
# -------------------------------------------------------------

def setup_internal_mails_folder(inbox_folder):
    """Creates (or reuses) a 'Internal_Mails' folder directly under Inbox."""
    return get_or_create_subfolder(inbox_folder, "Internal_Mails")




def find_all_date_folders(inbox_folder):
    """Returns EVERY subfolder directly under Inbox whose name looks like
    a date (YYYY-MM-DD) - i.e. every day-folder ever created, not just
    the last 7 days. Needed to find a conversation's ORIGINAL folder no
    matter how long ago it started."""
    date_folders = []
    for f in inbox_folder.Folders:
        if DATE_FOLDER_PATTERN.match(f.Name):
            date_folders.append(f)
    return date_folders

def delete_old_date_folders(inbox_folder, days):
    """Anchored to REAL current time (not DAYS_OFFSET) - cleanup should
    track the real calendar regardless of which day you're reprocessing."""
    cutoff = (datetime.now() - timedelta(days=days)).date()
    deleted = []
    for f in find_all_date_folders(inbox_folder):
        try:
            folder_date = datetime.strptime(f.Name, "%Y-%m-%d").date()
        except ValueError:
            continue
        if folder_date <= cutoff:
            age_days = (datetime.now().date() - folder_date).days
            print(f"Deleting folder '{f.Name}' ({age_days} days old)")
            f.Delete()
            deleted.append(f.Name)
    return deleted


def compute_episode_dates(full_messages):
    """
    Walks a conversation's full chronological message list (Customer +
    Support, already sorted by timestamp) and assigns each CUSTOMER
    message an "episode date" - the date of the folder it should file
    into.

    Rule:
      - Consecutive customer messages with NO support reply between them
        are one unanswered episode and share the date of the FIRST
        message in that run - so an ongoing, unanswered complaint keeps
        accumulating in its ORIGINAL folder (req: "no reply -> old date").
      - A support reply (including an actually-SENT resolve reply)
        closes the open episode. The next customer message - whether a
        plain follow-up or a reply to a resolved thread reopening it -
        starts a brand new episode dated to ITS OWN arrival day
        (req: "replied/resolved -> new date").
    """
    episode_date = None
    dates_by_entry_id = {}
    for m in full_messages:
        if m["direction"] == "Customer":
            if episode_date is None:
                episode_date = m["timestamp"].date()
            entry_id = m.get("entry_id")
            if entry_id:
                dates_by_entry_id[entry_id] = episode_date
        else:  # Support - closes whatever episode was open
            episode_date = None
    return dates_by_entry_id


def _strip_quoted_history(full_body):
    """
    Returns only the text BEFORE the first quoted-reply marker, so resolve
    detection only ever looks at what the customer actually just wrote -
    not our own earlier reply sitting in the quoted history below it.
    """
    new_text = full_body
    for marker in QUOTE_MARKERS:
        idx = new_text.find(marker)
        if idx != -1:
            new_text = new_text[:idx]
    return new_text


# Check whether the customer's message indicates the issue is resolved.
# category is the complaint category already computed by
# analyze_complaint_text() for this same message - reused here as a
# signal, not recomputed.
def check_resolve(msg, category="general"):
    full_body = msg.Body or ""
    if not full_body.strip():
        return False

    # Only check the customer's NEW text, not the quoted history below it.
    new_text = _strip_quoted_history(full_body)
    body = new_text.lower()
    if not body.strip():
        return False

    # A question mark almost always means there's something still open,
    # regardless of what resolution-sounding words appear elsewhere.
    if "?" in body:
        return False

    # "Thanks for the help, BUT I still have a question..." - a contrast
    # conjunction next to a resolved word means the customer is pivoting,
    # not closing out.
    if any(cw in body for cw in CONTRAST_WORDS):
        return False

    # Strong, unambiguous phrases are trusted on their own - nobody
    # writes "issue resolved" or "you may close the complaint" unless
    # they mean it.
    if any(phrase in body for phrase in STRONG_RESOLVED_PHRASES):
        return True 

    # Weak/ambiguous words ("thanks", "done", "fixed" used loosely) are
    # only trusted when there's no sign of a live, categorized complaint
    # riding alongside them. If the message also matches a real
    # complaint category (warranty, mechanical, etc.) it clearly isn't
    # just a courtesy close-out - e.g. "Thank you for your response, but
    # I want to ask about my car's warranty" matches "warranty" category
    # and must NOT be auto-closed off the word "thank you".
    has_weak_word = any(word in body for word in WEAK_RESOLVED_WORDS)
    if not has_weak_word:
        return False

    return category == "general"




def is_auto_resolvable(msg, category):
    """
    Wraps check_resolve() with a safety gate: safety-category
    conversations are never auto-closed off a keyword match at all -
    those need a human to confirm resolution regardless of wording.
    Every other category still gets check_resolve()'s own
    category-awareness (see WEAK_RESOLVED_WORDS logic above).
    """
    return check_resolve(msg, category)


def compute_status(has_number, is_resolved):
    """
    Determines the conversation's overall status:
        "-"       -> default, we don't even have the customer's number yet
        "Open"    -> we have their number and the complaint is still active
        "Resolve" -> the complaint has been marked resolved

    Resolve always takes priority - a conversation that's been resolved
    is reported as "Resolve" for this run's evaluation even if a number
    happens to be present, since resolution is the more significant state.
    """
    if is_resolved:
        return "Resolve"
    if has_number:
        return "Open"
    return "-"


# --------------------------------------------------------------------------
# PERSISTED "RESOLVED" STATE 
# --------------------------------------------------------------------------

def _has_resolved_tag(categories_str):
    """Category strings from Outlook are typically ';'-or-','-separated
    (e.g. 'Red Category; ResolvedConversation'). A plain substring check
    is enough here and is tolerant of either separator style."""
    return RESOLVED_CATEGORY_TAG in (categories_str or "")


def get_date_folder_cached(inbox_folder, date_obj, cache):
    """Returns (creating if needed) the date folder for date_obj, reusing a
    per-run cache (date_str -> folder) so the same date's folder is only
    looked up / created once even though many emails can share a date."""
    date_str = date_obj.strftime("%Y-%m-%d")
    if date_str not in cache:
        cache[date_str] = get_or_create_subfolder(inbox_folder, date_str)
    return cache[date_str]


def collect_inbox_dates(inbox_items):
    """Returns the sorted list of distinct dates found among the given
    Inbox mail items, skipping anything without a usable ReceivedTime."""
    dates = set()
    for m in inbox_items:
        try:
            if check_outlook_item(m.Class):
                continue
            received_time = m.ReceivedTime
            if received_time.tzinfo is not None:
                received_time = received_time.replace(tzinfo=None)
            dates.add(received_time.date())
        except Exception:
            continue
    return sorted(dates)


def ensure_date_folders_for_inbox(inbox_folder, inbox_items, cache):
    """Makes sure every date that currently has mail sitting in Inbox has a
    matching date folder created under Inbox, and returns the list of
    dates it ensured folders for."""
    dates = collect_inbox_dates(inbox_items)
    for d in dates:
        get_date_folder_cached(inbox_folder, d, cache)
    return dates


def mark_conversation_resolved(msg):
    """
    Persists the resolved state directly on the Outlook item's Categories
    field, so it survives across runs/days without any external storage -
    the same tagging pattern already used for escalation (subject tag)
    and follow-up nudges (Categories tag) elsewhere in this file.
    """
    try:
        existing = msg.Categories or ""
        if not _has_resolved_tag(existing):
            msg.Categories = f"{existing}; {RESOLVED_CATEGORY_TAG}" if existing else RESOLVED_CATEGORY_TAG
            msg.Save()
    except Exception as e:
        print(f"Could not tag conversation as resolved: {e}")


def build_persistently_resolved_conversations(customer_map):
    """
    Returns the set of ConversationIDs where ANY customer message ever
    filed away in a date folder carries the resolved tag. Once a
    conversation lands in this set it is treated as permanently closed
    for auto-drafting purposes - exactly like a real helpdesk ticket
    that stays "Solved" even if the customer's mail thread receives
    another message, until a human reopens it.
    """
    resolved_conv_ids = set()
    for conv_id, msgs in customer_map.items():
        for m in msgs:
            if _has_resolved_tag(m.get("categories", "")):
                resolved_conv_ids.add(conv_id)
                break
    return resolved_conv_ids


# --------------------------------------------------------------------------
# CACHED, SINGLE-PASS DATA COLLECTION
# (fix for the old per-conversation full-mailbox rescans)
# --------------------------------------------------------------------------

def build_sent_messages_map(outlook_namespace):
    """
    ONE pass over Sent Items for the whole run. Returns:
        conv_id -> sorted list of dicts (direction/sender/subject/body/timestamp)
    Used both for response timing (pairs) and for full escalation text
    trails, so Sent Items never needs to be rescanned per conversation.
    """
    sent_folder = outlook_namespace.GetDefaultFolder(5)  # 5 = olFolderSentMail
    conv_sent_map = {}
    for sent_msg in sent_folder.Items:
        try:
            if check_outlook_item(sent_msg.Class):
                continue
            conv_id = sent_msg.ConversationID
            if not conv_id:
                continue
            sent_on = sent_msg.SentOn
            if sent_on.tzinfo is not None:
                sent_on = sent_on.replace(tzinfo=None)
            conv_sent_map.setdefault(conv_id, []).append({
                "direction": "Support",
                "sender": "Support Team",
                "subject": sent_msg.Subject or "",
                "body": (sent_msg.Body or "").strip(),
                "timestamp": sent_on,
            })
        except Exception:
            continue

    for conv_id in conv_sent_map:
        conv_sent_map[conv_id].sort(key=lambda m: m["timestamp"])
    return conv_sent_map


def build_customer_messages_map(date_folders):
    """
    ONE pass over every existing date folder for the whole run. Returns:
        conv_map: conv_id -> list of customer-message dicts (full text,
            including the raw "categories" string so persisted state -
            e.g. the resolved tag - can be read back without a second scan)
        seen_entry_ids: set of every EntryID already accounted for, so
            today's still-in-Inbox messages can be added without
            double-counting anything already filed.
    """
    seen_entry_ids = set()
    conv_map = {}
    for f in date_folders:
        try:
            for m in f.Items:
                try:
                    if check_outlook_item(m.Class):
                        continue
                    conv_id = m.ConversationID
                    if not conv_id:
                        continue
                    entry_id = m.EntryID
                    if entry_id in seen_entry_ids:
                        continue
                    seen_entry_ids.add(entry_id)
                    received_time = m.ReceivedTime
                    if received_time.tzinfo is not None:
                        received_time = received_time.replace(tzinfo=None)
                    conv_map.setdefault(conv_id, []).append({
                        "direction": "Customer",
                        "sender": m.SenderName or get_sender_email(m),
                        "sender_email": get_sender_email(m),
                        "subject": m.Subject or "",
                        "body": (m.Body or "").strip(),
                        "timestamp": received_time,
                        "entry_id": entry_id,
                        "categories": m.Categories or "",
                    })
                except Exception:
                    continue
        except Exception:
            continue
    return conv_map, seen_entry_ids


def get_full_conversation_data(conv_id, customer_map, seen_entry_ids, sent_map, today_msgs):
    """
    Builds the complete (all-time) customer-message list AND the full
    chronological text trail for one conversation, using the prebuilt
    caches instead of rescanning any folder. `today_msgs` covers
    messages still sitting in the live Inbox that haven't been filed
    into a date folder yet this run.
    """
    customer_msgs = list(customer_map.get(conv_id, []))

    for m in today_msgs:
        try:
            entry_id = m.EntryID
            if entry_id in seen_entry_ids:
                continue
            received_time = m.ReceivedTime
            if received_time.tzinfo is not None:
                received_time = received_time.replace(tzinfo=None)
            customer_msgs.append({
                "direction": "Customer",
                "sender": m.SenderName or get_sender_email(m),
                "subject": m.Subject or "",
                "body": (m.Body or "").strip(),
                "timestamp": received_time,
                "entry_id": entry_id,
                "categories": m.Categories or "",
            })
        except Exception:
            continue

    support_msgs = sent_map.get(conv_id, [])

    full_messages = customer_msgs + support_msgs
    full_messages.sort(key=lambda m: m["timestamp"])

    full_receiving_times = [m["timestamp"] for m in customer_msgs]

    return full_receiving_times, full_messages


# _____________ NUM CHECKER ___________________________________________________________________________

PAKISTANI_NUMBER_PATTERN = re.compile(r'^(0|\+92|0092)3\d{9}$')


def check_num(subject, body):
    """
    Returns True only if subject or body contains a validly-formatted
    Pakistani mobile number:
        - Local:         03XXXXXXXXX     (11 digits total, starts with 03)
        - International: +923XXXXXXXXX   (+92 followed by 3 + 9 digits)
        - International: 00923XXXXXXXXX  (0092 followed by 3 + 9 digits)
    Any other digit sequence (dates, ticket numbers, incomplete/too-short
    or too-long numbers) is correctly rejected.
    """
    text = f"{subject or ''} {body or ''}"

    candidates = re.split(r'[^\d+()\s.-]+', text)

    for candidate in candidates:
        cleaned = re.sub(r'[\s().-]', '', candidate)
        if PAKISTANI_NUMBER_PATTERN.match(cleaned):
            return True

    return False


def check_num_in_history(customer_messages):
    """
    customer_messages: list of dicts with 'subject' and 'body' keys,
    OR list of raw Outlook message objects (has .Subject/.Body).
    Returns True if ANY customer message in the conversation contains
    a valid number, not just the latest one.
    """
    for m in customer_messages:
        if isinstance(m, dict):
            subj, body = m.get("subject", ""), m.get("body", "")
        else:
            subj, body = m.Subject or "", m.Body or ""
        if check_num(subj, body):
            return True
    return False
# -----------------------------------------------------------------------------------------------------------------


# Generate Suggested reply by name and based on category
def generate_suggested_reply(category, sender_name):
    template = REPLY_TEMPLATES.get(category, REPLY_TEMPLATES["general"])
    first_name = sender_name.split()[0] if sender_name else "Customer"
    return template.format(name=first_name)


# Get ids of already created Drafts that are still in drafts folder
def build_existing_draft_conversations(outlook_namespace):
    drafts_folder = outlook_namespace.GetDefaultFolder(16)  # 16 = olFolderDrafts
    conv_ids = set()
    for item in drafts_folder.Items:
        try:
            #Drraft must be an email
            if not check_outlook_item(item.Class) and item.ConversationID:
                conv_ids.add(item.ConversationID)
        except Exception:
            continue
    return conv_ids


def build_followup_sent_map(outlook_namespace):
    """
    Returns dict: conv_id -> set of support-reply-timestamp strings that
    have ALREADY had a follow-up nudge sent for them - tracked via
    Categories (invisible to the customer, only visible in our own Sent
    Items). If our team replies AGAIN after a nudge and the customer goes
    silent once more, the timestamp changes - so a new nudge becomes
    eligible per silence streak, rather than firing only once ever.
    """
    sent_folder = outlook_namespace.GetDefaultFolder(5)  # 5 = olFolderSentMail
    followup_map = {}
    for item in sent_folder.Items:
        try:
            if check_outlook_item(item.Class):
                continue
            categories = item.Categories or ""
            if categories.startswith("FollowupSent|"):
                parts = categories.split("|")
                if len(parts) == 3:
                    _, conv_id, ts = parts
                    followup_map.setdefault(conv_id, set()).add(ts)
        except Exception:
            continue
    return followup_map


def send_followup_to_customer(conv_id, customer_email, customer_name, subject, last_support_reply_time, outlook_app):
    """
    Sends a gentle follow-up nudge DIRECTLY TO THE CUSTOMER when they've
    gone silent for FOLLOWUP_DAYS_THRESHOLD+ days after our team's last
    reply, and the complaint isn't marked resolved. The tracking tag is
    stored in Categories, NOT the visible subject - the customer only
    ever sees a clean "Following up: <subject>" line.
    """
    template = REPLY_TEMPLATES["Followup_Nudge"]
    first_name = customer_name.split()[0] if customer_name else "Customer"
    body_text = template.format(name=first_name)
    ts_tag = last_support_reply_time.strftime("%Y%m%d%H%M%S")

    try:
        mail = outlook_app.CreateItem(0)  # 0 = olMailItem
        mail.To = customer_email
        mail.Subject = f"Following up: {subject}"
        mail.HTMLBody = f"<p>{body_text.replace(chr(10), '<br>')}</p>"
        mail.Categories = f"FollowupSent|{conv_id}|{ts_tag}"
        mail.Send()
        print(f"Follow-up nudge sent to {customer_email} for '{subject}'")
        return True
    except Exception as e:
        print(f"Could not send follow-up nudge for '{subject}': {e}")
        return False


def build_escalated_conversations(outlook_namespace):
    """
    Returns the set of CUSTOMER ConversationIDs that have already been
    successfully escalated to the HOD (i.e. an escalation email with
    that tag genuinely exists in Sent Items - see fix #3 in the
    changelog for why this matters).
    """
    sent_folder = outlook_namespace.GetDefaultFolder(5)  # Sent Items
    escalated_ids = set()
    pattern = re.compile(r"\[HOD_ESCALATED\]\[(.*?)\]")

    for item in sent_folder.Items:
        try:
            if check_outlook_item(item.Class):
                continue
            subject = item.Subject or ""
            match = pattern.search(subject)
            if match:
                escalated_ids.add(match.group(1))
        except Exception:
            continue

    return escalated_ids

def send_hod_overdue_popup(conv_id, subject, sender_name, sender_email,
                            received_time, business_days_waited, outlook_app):
    """Sends an urgent alert to HOD_EMAIL flagging a customer message
    that has gone unanswered for business_days_waited working day(s)."""
    ts_tag = received_time.strftime("%Y%m%d%H%M%S")
    try:
        mail = outlook_app.CreateItem(0)
        mail.To = HOD_EMAIL
        mail.Subject = f"[HOD_OVERDUE][{conv_id}] {business_days_waited}+ working days unanswered - {subject}"
        mail.HTMLBody = (
            f"<p style='color:#C00000;font-weight:bold;'>"
            f"ALERT: No reply sent for {business_days_waited} working day(s).</p>"
            f"<ul>"
            f"<li><b>Customer:</b> {sender_name} ({sender_email})</li>"
            f"<li><b>Subject:</b> {subject}</li>"
            f"<li><b>Received:</b> {received_time.strftime('%Y-%m-%d %H:%M:%S')}</li>"
            f"</ul>"
        )
        mail.Categories = f"{HOD_POPUP_TAG_PREFIX}{conv_id}|{ts_tag}"
        mail.Send()
        print(f"HOD overdue alert sent for '{subject}' ({business_days_waited} working days)")
        return True
    except Exception as e:
        print(f"Could not send HOD overdue alert for '{subject}': {e}")
        return False


def send_hod_escalation_email(customer_conv_id, subject, sender_name,
                               sender_email, full_pairs,
                               full_messages, outlook_app):
    """
    Sends escalation email to HOD.
    Returns True only if mail.Send() actually completed without raising -
    callers must NOT mark the conversation as escalated on a False
    return, or a failed send permanently (and silently) suppresses a
    real escalation.
    """
    total_received = len(full_pairs)
    replied_count = sum(1 for _, response in full_pairs if response != "-")

    message_blocks = ""
    for i, m in enumerate(full_messages, start=1):
        body_text = m["body"][:3000]
        truncated_note = " <i>(truncated)</i>" if len(m["body"]) > 3000 else ""
        color = "#4472C4" if m["direction"] == "Customer" else "#70AD47"

        message_blocks += (
            f"<div style='border-left:4px solid {color};padding:8px 12px;margin-bottom:12px;'>"
            f"<b>#{i} - {m['direction']}</b> ({m['sender']}) "
            f"- {m['timestamp'].strftime('%Y-%m-%d %H:%M:%S')}<br>"
            f"<b>Subject:</b> {m['subject']}<br>"
            f"<b>Message:</b>{truncated_note}<br>"
            f"{body_text.replace(chr(10), '<br>')}"
            f"</div>"
        )

    try:
        mail = outlook_app.CreateItem(0)
        mail.To = HOD_EMAIL
        mail.Subject = f"[HOD_ESCALATED][{customer_conv_id}] {subject}"
        mail.HTMLBody = (
            f"<p>This conversation has exceeded "
            f"{ESCALATION_THRESHOLD} exchanges.</p>"
            f"<ul>"
            f"<li><b>Customer:</b> {sender_name} ({sender_email})</li>"
            f"<li><b>Subject:</b> {subject}</li>"
            f"<li><b>Total Received:</b> {total_received}</li>"
            f"<li><b>Total Replied:</b> {replied_count}</li>"
            f"</ul>"
            f"<p><b>Complete Message Trail:</b></p>"
            f"{message_blocks}"
        )
        mail.Send()
        print(f"Escalation email sent for {customer_conv_id}")
        return True
    except Exception as e:
        print(f"FAILED to send escalation email for {customer_conv_id}: {e}")
        return False


def create_draft_reply(msg, suggested_reply, existing_draft_conversations):
    if not IS_CUSTOMER_CARE:
        return
    try:
        conv_id = msg.ConversationID
        if conv_id and conv_id in existing_draft_conversations:
            print(f"Draft already exists for '{msg.Subject}' - skipping duplicate.")
            return
        reply = msg.Reply()
        formatted_reply = suggested_reply.replace("\n", "<br>")
        reply.HTMLBody = f"<p>{formatted_reply}</p><hr>" + reply.HTMLBody
        reply.Save()
        print(f"Draft created for '{msg.Subject}'")
        if conv_id:
            existing_draft_conversations.add(conv_id)
    except Exception as e:
        print(f"Could not create draft for '{msg.Subject}': {e}")


# Analyze complaint text from URGENT KEYWORDS and category keywords, used to build a priority
def analyze_complaint_text(subject, body):
    """Scans subject+body ONCE and returns (category, has_urgent_keyword).
    Category and urgency are scored independently (see CATEGORY_KEYWORDS
    changelog note) so an urgent-but-unrelated email is never
    misclassified as "safety" just for containing "ASAP"."""
    text = f"{subject or ''} {body or ''}".lower()
    has_urgent_keyword = any(kw in text for kw in URGENT_KEYWORDS)
    for category in ("safety", "mechanical", "warranty", "service_delay"):
        for kw in CATEGORY_KEYWORDS[category]:
            if kw in text:
                return category, has_urgent_keyword
    return "general", has_urgent_keyword


# Compute Priority
def compute_priority(category, has_urgent_keyword, still_unanswered):
    """
    Simple priority: safety complaints and urgent language are always
    High. Otherwise Medium for mechanical/warranty/service issues still
    awaiting a reply, Low for everything else (including anything already
    answered).
    """
    if category == "safety" or has_urgent_keyword:
        return "High"
    if still_unanswered and category in ("mechanical", "warranty", "service_delay"):
        return "Medium"
    return "Low"


# Move msg to target_folder like to today date folder/Promotions etc
def move_to_target_folder(msg, target_folder):
    """Moves the email into the given date folder."""
    try:
        msg.Move(target_folder)
    except Exception as e:
        print(f"Could not move '{msg.Subject}' to '{target_folder.Name}': {e}")

#Move to target Folder
def move_to_internal_folder(msg, internal_folder):
    """Moves the email into the Internal_Mails folder (removes it from wherever it currently is)."""
    try:
        msg.Move(internal_folder)
        print(f"Moved '{msg.Subject}' to '{internal_folder.Name}'")
    except Exception as e:
        print(f"Could not move '{msg.Subject}': {e}")

def build_conversation_pairs(receiving_times, response_times_for_conv):
    receiving_times = sorted(receiving_times)
    response_times_for_conv = sorted(response_times_for_conv)

    responses = ["-"] * len(receiving_times)
    used = [False] * len(response_times_for_conv)

    # Match from newest customer message backwards
    for i in range(len(receiving_times) - 1, -1, -1):
        r_time = receiving_times[i]
        for j in range(len(response_times_for_conv) - 1, -1, -1):
            if used[j]:
                continue
            if response_times_for_conv[j] >= r_time:
                responses[i] = response_times_for_conv[j].strftime("%Y-%m-%d %H:%M:%S")
                used[j] = True
                break

    pairs = []
    for r_time, resp in zip(receiving_times, responses):
        pairs.append((r_time.strftime("%Y-%m-%d %H:%M:%S"), resp))

    return pairs


def calculate_avg_response_time(pairs):
    """
    Given the (receiving_time_str, response_time_str) pairs for a
    conversation, returns the average response time as a readable
    string (e.g. "2h 15m"), using only the pairs that have an actual
    reply. Returns "-" if none have been replied to yet.
    """
    deltas = []
    for receiving_str, response_str in pairs:
        if response_str == "-":
            continue
        r_time = datetime.strptime(receiving_str, "%Y-%m-%d %H:%M:%S")
        s_time = datetime.strptime(response_str, "%Y-%m-%d %H:%M:%S")
        deltas.append((s_time - r_time).total_seconds())

    if not deltas:
        return "-"

    avg_seconds = sum(deltas) / len(deltas)
    hours = int(avg_seconds // 3600)
    minutes = int((avg_seconds % 3600) // 60)
    if hours > 0:
        return f"{hours}h {minutes}m"
    return f"{minutes}m"


def count_consecutive_unanswered(pairs):
    """
    Counts how many messages in a row, counting backwards from the most
    recent, have no response yet. Returns "-" if the streak is only 1
    (that's just "waiting for a first reply", not a repeated follow-up
    pattern) - only returns an actual number when the streak is 2+.
    """
    streak = 0
    for _, response in reversed(pairs):
        if response == "-":
            streak += 1
        else:
            break
    if streak == 0:
        return "-"
    return streak


def _safe_mail_items(items):
    """Returns only real MailItems (Class 43) that actually expose
    ReceivedTime - Outlook Inbox/date folders can contain meeting
    requests, receipts, and reports that don't have this property,
    and those must never reach a sort/filter step that assumes they do."""
    safe = []
    for it in items:
        try:
            if check_outlook_item(it.Class):
                continue
            _ = it.ReceivedTime  # touch it now, inside a try, not later in sorted()
            safe.append(it)
        except Exception:
            continue
    return safe


def extract_conversations(folder, date_folder_cache, all_date_folders, sent_map,
                           customer_map, seen_entry_ids, escalated_conversations,
                           existing_draft_conversations, limit=None):
    """
    ...
    """
    outlook_app = folder.Application

    #conversation_folder_map = build_conversation_folder_map(all_date_folders)          -------------NOT use anywhere

    # Conversations that were EVER marked "Resolve" in a previous run stay
    # closed for auto-drafting purposes (fix #10), no matter what a later
    # message in the same thread says.
    persistently_resolved_conversations = build_persistently_resolved_conversations(customer_map)
    target_date = get_target_date().date()

    # inbox_items = list(folder.Items)
    # inbox_items_sorted = sorted(inbox_items, key=lambda m: m.ReceivedTime, reverse=True)
    # already_moved_items = list(date_folder.Items)

    inbox_items = _safe_mail_items(list(folder.Items))

    # create/reuse a date folder for EVERY distinct date currently in Inbox
    relevant_dates = ensure_date_folders_for_inbox(folder, inbox_items, date_folder_cache)

    inbox_items_sorted = sorted(inbox_items, key=lambda m: m.ReceivedTime, reverse=True)

    already_moved_items = []
    for d in relevant_dates:
        fdr = date_folder_cache.get(d.strftime("%Y-%m-%d"))
        if fdr is not None:
            already_moved_items.extend(_safe_mail_items(list(fdr.Items)))

    combined = [(m, True) for m in inbox_items_sorted] + [(m, False) for m in already_moved_items]

    conversations = {}  # conv_id -> {"messages": [(msg, needs_move), ...], "receiving_times": [...]}
    processed_count = 0

    for msg, needs_move in combined:
        try:
            if check_outlook_item(msg.Class):
                continue
            if not is_primary(msg):
                continue

            received_time = msg.ReceivedTime
            if received_time.tzinfo is not None:
                received_time = received_time.replace(tzinfo=None)
            if not PROCESS_ALL_INBOX_DATES and received_time.date() != target_date:
                continue

            if needs_move and check_promotion(msg.Subject, msg.Body):
                print(f"Moving promotional email to Promotions: '{msg.Subject}'")
                promotion_folder = setup_promotions_folder(folder)
                move_to_target_folder(msg, promotion_folder)
                continue

            sender = get_sender_email(msg).lower()
            if "noreply" in sender or "no-reply" in sender:
                continue

            conv_id = msg.ConversationID
            if not conv_id:
                continue

            conversations.setdefault(conv_id, {"messages": [], "receiving_times": []})
            conversations[conv_id]["messages"].append((msg, needs_move))
            conversations[conv_id]["receiving_times"].append(received_time)

            processed_count += 1
            if limit and processed_count >= limit:
                break
        except Exception as e:
            print(f"Skipped one email due to error: {e}")
            continue

    rows = []
    for conv_id, data in conversations.items():
        msgs = [m for m, _ in data["messages"]]
        receiving_times = data["receiving_times"]

        latest_msg = max(msgs, key=lambda m: m.ReceivedTime)
        subject = latest_msg.Subject or ""
        body = latest_msg.Body or ""
        category, has_urgent_keyword = analyze_complaint_text(subject, body)

        response_times_for_conv = [m["timestamp"] for m in sent_map.get(conv_id, [])]
        pairs = build_conversation_pairs(receiving_times, response_times_for_conv)
        still_unanswered = pairs[-1][1] == "-" if pairs else True
        full_receiving_times, full_messages = get_full_conversation_data(
        conv_id, customer_map, seen_entry_ids, sent_map, msgs
        )
        episode_dates = compute_episode_dates(full_messages)

        latest_customer_time = None
        latest_our_reply_time = None
        for m in full_messages:
            if m["direction"] == "Customer":
                if latest_customer_time is None or m["timestamp"] > latest_customer_time:
                    latest_customer_time = m["timestamp"]
            elif m["direction"] == "Support":
                if latest_our_reply_time is None or m["timestamp"] > latest_our_reply_time:
                    latest_our_reply_time = m["timestamp"]

        already_replied = (
            latest_customer_time is not None
            and latest_our_reply_time is not None
            and latest_our_reply_time >= latest_customer_time
        )

        # A conversation that was resolved in the past stays resolved for
        # status/drafting purposes even if today's message alone wouldn't
        # trigger the keyword-based detector.
        conv_already_resolved = conv_id in persistently_resolved_conversations

        # Fix #5: resolve detection is gated off for safety complaints.
        is_resolved = conv_already_resolved or is_auto_resolvable(latest_msg, category)
        customer_msgs_only = [m for m in full_messages if m["direction"] == "Customer"]
        has_number = check_num_in_history(customer_msgs_only)
        status = compute_status(has_number, is_resolved)

        # Fix #2 + Fix #10: draft-reply logic is mutually exclusive and
        # ordered resolve -> missing number -> normal reply, EXCEPT that a
        # conversation already persisted as resolved never gets ANY new
        # auto-draft at all - a human has to reopen/handle it manually,
        # the same way a closed ticket doesn't keep auto-replying just
        # because the customer's thread got one more message.
        # if conv_already_resolved:
        #     #print(f"Skipping auto-draft for '{subject}' - conversation already marked Resolve.")
        #     print()

        if is_resolved and not conv_already_resolved and IS_CUSTOMER_CARE:
            suggested_reply = generate_suggested_reply("Complain_Resolve", latest_msg.SenderName)
            create_draft_reply(latest_msg, suggested_reply, existing_draft_conversations)
            mark_conversation_resolved(latest_msg)

        
        if not already_replied and not status == "Resolve" and IS_CUSTOMER_CARE:
            if not has_number :
                suggested_reply = generate_suggested_reply("Missing_Num", latest_msg.SenderName)
                create_draft_reply(latest_msg, suggested_reply, existing_draft_conversations)
            
            elif still_unanswered:
                suggested_reply = generate_suggested_reply(category, latest_msg.SenderName)
                create_draft_reply(latest_msg, suggested_reply, existing_draft_conversations)

        total_received = len(pairs)
        replied_count = sum(1 for _, response in pairs if response != "-")
        unreplied_count = total_received - replied_count

        priority = compute_priority(category, has_urgent_keyword, still_unanswered)

        full_pairs = build_conversation_pairs(full_receiving_times, response_times_for_conv)
        full_total_received = len(full_pairs)
        full_replied_count = sum(1 for _, response in full_pairs if response != "-")
        consecutive_unanswered = count_consecutive_unanswered(full_pairs)
        unresolved_backlog = full_total_received - full_replied_count

        is_sent = False
        already_escalated = conv_id in escalated_conversations


        # Trigger A: 5+ consecutive unanswered follow-ups in a row.
        if (isinstance(consecutive_unanswered, int) 
            and consecutive_unanswered >= 5 
            and status == "Open" 
            and not already_escalated
            and IS_CUSTOMER_CARE):
            sent_ok = send_hod_escalation_email(
                conv_id, subject, latest_msg.SenderName,
                get_sender_email(latest_msg), full_pairs, full_messages, outlook_app
            )
            if sent_ok:  # fix #3: only mark escalated if the email really sent
                escalated_conversations.add(conv_id)
                is_sent = True

        # Trigger B (fix #4): a long conversation with a genuine unresolved
        # backlog (2+ messages never replied to), NOT a long-but-healthy
        # fully-answered thread like the original "replied_count >= 5" check.
        if (full_total_received >= ESCALATION_THRESHOLD
                and full_replied_count >= ESCALATION_THRESHOLD
                and IS_CUSTOMER_CARE
                and status == "Open"
                and not already_escalated
                and not is_sent
                and IS_CUSTOMER_CARE):

            sent_ok = send_hod_escalation_email(
                conv_id, subject, latest_msg.SenderName,
                get_sender_email(latest_msg), full_pairs, full_messages, outlook_app
            )
            if sent_ok:
                escalated_conversations.add(conv_id)

        row = {
            "Message ID": get_message_id(latest_msg),
            "Subject": subject,
            "Sender Name": latest_msg.SenderName or "",
            "Sender Email": get_sender_email(latest_msg),
            "Category": category,
            "Priority": priority,
            "Status": status,
            "Total Received": total_received,
            "Replied": replied_count,
            "Unreplied Count": unreplied_count,
            "Consecutive Unanswered": consecutive_unanswered,
            "Avg Response Time": calculate_avg_response_time(pairs),
            "Is Resolved": is_resolved,
            "_pairs": pairs,
        }
        rows.append(row)

        for msg, needs_move in data["messages"]:
            if needs_move:
                entry_id = msg.EntryID
                ep_date = episode_dates.get(entry_id)
                if ep_date is None:  # fallback safety net, shouldn't normally hit
                    rt = msg.ReceivedTime
                    ep_date = (rt.replace(tzinfo=None) if rt.tzinfo else rt).date()
                target_folder = get_date_folder_cached(folder, ep_date, date_folder_cache)
                move_to_target_folder(msg, target_folder)

    return rows


def write_to_excel(rows, output_path="emails_export.xlsx"):
    """Writes one row per conversation, with dynamically-expanded
    Receiving Time N / Response Time N column pairs based on whichever
    conversation had the most exchanges today."""
    if not rows:
        print("No conversations to write.")
        return

    priority_order = {"High": 0, "Medium": 1, "Low": 2}
    rows = sorted(rows, key=lambda r: priority_order.get(r.get("Priority", "Low"), 3))

    max_pairs = max(len(r["_pairs"]) for r in rows)

    fixed_headers = ["Message ID", "Subject", "Sender Name", "Sender Email", "Category", "Priority", "Status",
                      "Total Received", "Replied", "Unreplied Count", "Consecutive Unanswered",
                      "Avg Response Time", "Is Resolved"]
    pair_headers = []
    for i in range(1, max_pairs + 1):
        pair_headers.append(f"Receiving Time {i}")
        pair_headers.append(f"Response Time {i}")
    headers = fixed_headers + pair_headers

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Emails"
    ws.append(headers)

    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill

    priority_fills = {
        "High": openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "Medium": openpyxl.styles.PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    }
    status_fills = {
        "Open": openpyxl.styles.PatternFill(
            start_color="FFC7CE",  # Light Red
            end_color="FFC7CE",
            fill_type="solid"
        ),
        "Resolve": openpyxl.styles.PatternFill(
            start_color="C6EFCE",  # Light Green
            end_color="C6EFCE",
            fill_type="solid"
        ),
    }
    priority_col_index = headers.index("Priority") + 1

    for row_num, row in enumerate(rows, start=2):
        row_values = [row.get(h, "") for h in fixed_headers]
        for i in range(1, max_pairs + 1):
            if i <= len(row["_pairs"]):
                receiving, response = row["_pairs"][i - 1]
            else:
                receiving, response = "-", "-"
            row_values.append(receiving)
            row_values.append(response)
        ws.append(row_values)

        fill = priority_fills.get(row.get("Priority", ""))
        if fill:
            ws.cell(row=row_num, column=priority_col_index).fill = fill

        status_col_index = headers.index("Status") + 1

        status_fill = status_fills.get(row.get("Status", ""))
        if status_fill:
            ws.cell(row=row_num, column=status_col_index).fill = status_fill

    for col_num, header in enumerate(headers, 1):
        max_length = max(
            [len(str(header))] + [len(str(ws.cell(row=r, column=col_num).value or "")) for r in range(2, len(rows) + 2)]
        )
        ws.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 40)

    wb.save(output_path)
    print(f"Exported {len(rows)} conversations to '{output_path}' ({max_pairs} exchange pair(s) wide)")



def find_date_folders_last_n_days(inbox_folder, days):
    """Window anchored to target date (shifts with DAYS_OFFSET), looking
    back `days` days from there - `days` is now actually respected."""
    cutoff = (get_target_date() - timedelta(days=days - 1)).date()
    matched = []
    for f in inbox_folder.Folders:
        try:
            folder_date = datetime.strptime(f.Name, "%Y-%m-%d").date()
        except ValueError:
            continue
        if folder_date >= cutoff:
            matched.append(f)
    return matched

def extract_conversations_last_n_days(inbox_folder, days, sent_map, limit=None):
    date_folders = find_date_folders_last_n_days(inbox_folder, days=days)
    if not date_folders:
        print(f"No date folders found in the last {days} days.")
        return []

    conversations = {}
    processed_count = 0

    for date_folder in date_folders:
        for msg in date_folder.Items:
            try:
                if check_outlook_item(msg.Class):
                    continue
                if not is_primary(msg):
                    continue
                sender = get_sender_email(msg).lower()
                if "noreply" in sender or "no-reply" in sender:
                    continue
                conv_id = msg.ConversationID
                if not conv_id:
                    continue
                received_time = msg.ReceivedTime
                if received_time.tzinfo is not None:
                    received_time = received_time.replace(tzinfo=None)
                conversations.setdefault(conv_id, {"messages": [], "receiving_times": []})
                conversations[conv_id]["messages"].append(msg)
                conversations[conv_id]["receiving_times"].append(received_time)
                processed_count += 1
                if limit and processed_count >= limit:
                    break
            except Exception as e:
                print(f"Skipped one email due to error: {e}")
                continue

    rows = []
    for conv_id, data in conversations.items():
        msgs = data["messages"]
        receiving_times = data["receiving_times"]

        latest_msg = max(msgs, key=lambda m: m.ReceivedTime)
        subject = latest_msg.Subject or ""
        body = latest_msg.Body or ""
        category, has_urgent_keyword = analyze_complaint_text(subject, body)

        response_times_for_conv = [m["timestamp"] for m in sent_map.get(conv_id, [])]
        pairs = build_conversation_pairs(receiving_times, response_times_for_conv)
        still_unanswered = pairs[-1][1] == "-" if pairs else True

        total_received = len(pairs)
        replied_count = sum(1 for _, response in pairs if response != "-")
        consecutive_unanswered = count_consecutive_unanswered(pairs)
        unreplied_count = total_received - replied_count

        priority = compute_priority(category, has_urgent_keyword, still_unanswered)
        is_resolved = is_auto_resolvable(latest_msg, category)
        has_number = check_num_in_history(msgs)
        status = compute_status(has_number, is_resolved)

        row = {
            "Message ID": get_message_id(latest_msg),
            "Subject": subject,
            "Sender Name": latest_msg.SenderName or "",
            "Sender Email": get_sender_email(latest_msg),
            "Category": category,
            "Priority": priority,
            "Status": status,
            "Total Received": total_received,
            "Replied": replied_count,
            "Unreplied Count": unreplied_count,
            "Consecutive Unanswered": consecutive_unanswered,
            "Avg Response Time": calculate_avg_response_time(pairs),
            "Is Resolved": is_resolved,
            "_pairs": pairs,
        }
        rows.append(row)

    return rows


def check_and_process_internal_mails(folder, sent_map):

    internal_folder = setup_internal_mails_folder(folder)
    allowed_domains_lower = [d.lower().lstrip("@") for d in INTERNAL_MAIL_DOMAINS]

    # No date window - scan every date folder that exists, plus Inbox.
    date_folders = find_all_date_folders(folder)

    inbox_items = list(folder.Items)
    combined = [(m, True) for m in inbox_items]
    for df in date_folders:
        for m in df.Items:
            combined.append((m, False))

    conversations = {}
    for msg, needs_move in combined:
        try:
            if check_outlook_item(msg.Class):
                continue
            if not is_primary(msg):
                continue
            received_time = msg.ReceivedTime
            if received_time.tzinfo is not None:
                received_time = received_time.replace(tzinfo=None)
            sender = get_sender_email(msg)
            if not is_internal_domain(sender, allowed_domains_lower):
                continue
            conv_id = msg.ConversationID
            if not conv_id:
                continue
            conversations.setdefault(conv_id, {"messages": [], "receiving_times": []})
            conversations[conv_id]["messages"].append((msg, needs_move))
            conversations[conv_id]["receiving_times"].append(received_time)
        except Exception as e:
            print(f"Skipped one email while scanning internal mails: {e}")
            continue

    rows = []
    for conv_id, data in conversations.items():
        msgs = [m for m, _ in data["messages"]]
        receiving_times = data["receiving_times"]

        latest_msg = max(msgs, key=lambda m: m.ReceivedTime)
        subject = latest_msg.Subject or ""
        body = latest_msg.Body or ""
        category, has_urgent_keyword = analyze_complaint_text(subject, body)

        response_times_for_conv = [m["timestamp"] for m in sent_map.get(conv_id, [])]
        pairs = build_conversation_pairs(receiving_times, response_times_for_conv)
        still_unanswered = pairs[-1][1] == "-" if pairs else True

        total_received = len(pairs)
        replied_count = sum(1 for _, response in pairs if response != "-")
        unreplied_count = total_received - replied_count
        consecutive_unanswered = count_consecutive_unanswered(pairs)

        priority = compute_priority(category, has_urgent_keyword, still_unanswered)
        is_resolved = is_auto_resolvable(latest_msg, category)
        has_number = check_num_in_history(msgs)
        status = compute_status(has_number, is_resolved)

        row = {
            "Message ID": get_message_id(latest_msg),
            "Subject": subject,
            "Sender Name": latest_msg.SenderName or "",
            "Sender Email": get_sender_email(latest_msg),
            "Category": category,
            "Priority": priority,
            "Status": status,
            "Total Received": total_received,
            "Replied": replied_count,
            "Unreplied Count": unreplied_count,
            "Consecutive Unanswered": consecutive_unanswered,
            "Avg Response Time": calculate_avg_response_time(pairs),
            "Is Resolved": is_resolved,
            "_pairs": pairs,
        }
        rows.append(row)

        for msg, _needs_move in data["messages"]:
            move_to_internal_folder(msg, internal_folder)

    print(f"Internal mails processed: {len(rows)} conversation(s) copied to 'Internal_Mails'")
    return rows


def check_and_notify_hod_overdue(folder, customer_map, sent_map, hod_popup_sent_map,
                                  business_days_threshold=None):
    """For every conversation, finds the latest customer message. If our
    team hasn't replied to it yet AND it's been sitting unanswered for
    HOD_POPUP_BUSINESS_DAYS+ working days (Mon-Fri), alerts HOD - once
    per unanswered message, not once per run."""
    if business_days_threshold is None:
        business_days_threshold = HOD_POPUP_BUSINESS_DAYS

    outlook_app = folder.Application
    now = datetime.now()

    sent_count = 0
    for conv_id, customer_msgs in customer_map.items():
        if not customer_msgs:
            continue

        latest_customer_msg = max(customer_msgs, key=lambda m: m["timestamp"])
        support_msgs = sent_map.get(conv_id, [])
        latest_support_msg = max(support_msgs, key=lambda m: m["timestamp"]) if support_msgs else None

        # Already replied after this customer message? not overdue.
        if latest_support_msg and latest_support_msg["timestamp"] >= latest_customer_msg["timestamp"]:
            continue

        overdue_threshold = add_business_days(latest_customer_msg["timestamp"], business_days_threshold)
        if now < overdue_threshold:
            continue
        business_days_waited = business_days_threshold  # for the email/log message

        ts_tag = latest_customer_msg["timestamp"].strftime("%Y%m%d%H%M%S")
        if ts_tag in hod_popup_sent_map.get(conv_id, set()):
            continue  # already alerted for this exact unanswered message

        sent_ok = send_hod_overdue_popup(
            conv_id, latest_customer_msg["subject"], latest_customer_msg["sender"],
            latest_customer_msg.get("sender_email", ""), latest_customer_msg["timestamp"],
            business_days_waited, outlook_app
        )
        if sent_ok:
            sent_count += 1

    print(f"HOD overdue alerts sent: {sent_count}")
    return sent_count


def check_and_send_followup_nudges(folder, customer_map, sent_map, followup_sent_map,
                                    followup_days_threshold=None):
    """
    ...
    """
    if followup_days_threshold is None:
        followup_days_threshold = FOLLOWUP_DAYS_THRESHOLD

    outlook_app = folder.Application
    now = datetime.now()

    persistently_resolved_conversations = build_persistently_resolved_conversations(customer_map)
    class _MsgShim:
        """Tiny stand-in so is_auto_resolvable() can read .Body without a real COM object."""
        def __init__(self, body):
            self.Body = body

    sent_count = 0
    for conv_id, customer_msgs in customer_map.items():
        if not customer_msgs:
            continue

        latest_customer_msg = max(customer_msgs, key=lambda m: m["timestamp"])
        support_msgs = sent_map.get(conv_id, [])
        if not support_msgs:
            continue  # we never replied at all - different problem, not this one

        latest_support_msg = max(support_msgs, key=lambda m: m["timestamp"])

        if latest_support_msg["timestamp"] <= latest_customer_msg["timestamp"]:
            continue  # customer already wrote back after our reply

        conv_already_resolved = conv_id in persistently_resolved_conversations
        category, _ = analyze_complaint_text(latest_customer_msg["subject"], latest_customer_msg["body"])
        is_resolved = conv_already_resolved or is_auto_resolvable(_MsgShim(latest_customer_msg["body"]), category)
        has_number = check_num_in_history(customer_msgs)  # customer_msgs = full history, already in scope
        status = compute_status(has_number, is_resolved)

        if status != "Open":
            continue  # only nudge conversations that are genuinely still open

        days_silent = (now - latest_support_msg["timestamp"]).days
        if days_silent < followup_days_threshold:
            continue

        ts_tag = latest_support_msg["timestamp"].strftime("%Y%m%d%H%M%S")
        if ts_tag in followup_sent_map.get(conv_id, set()):
            continue  # already nudged for this exact silence streak

        customer_email = latest_customer_msg.get("sender_email", "")
        if not customer_email or "@" not in customer_email:
            print(f"Skipping follow-up for {conv_id}: no valid customer email found.")
            continue

        sent_ok = send_followup_to_customer(
            conv_id, customer_email, latest_customer_msg["sender"],
            latest_customer_msg["subject"], latest_support_msg["timestamp"], outlook_app
        )
        if sent_ok:
            sent_count += 1

    print(f"Follow-up nudges sent: {sent_count}")
    return sent_count


if __name__ == "__main__":
    move_all_spam_to_inbox()
    safe_email = (
        PERSONAL_EMAIL
        .replace("@", "_at_")
        .replace(".", "_")
    )
    target_date = get_target_date().date()
    target_date_str = target_date.strftime("%Y%m%d")

    refresh_outlook()  # fix #7: only runs when this file is executed directly

    LIMIT = None
    OUTPUT_FILE = os.path.join(
        PROJECT_DIRECTORY,
        f"emails_export_{target_date_str}.xlsx"
    )


    folder = get_outlook_folder(OUTLOOK_FOLDER_NAME)
    date_folder_cache = {}  # date_str -> folder, populated per-run
    outlook_namespace = folder.Application.GetNamespace("MAPI")
    all_date_folders = find_all_date_folders(folder)
    sent_map = build_sent_messages_map(outlook_namespace)
    customer_map, seen_entry_ids = build_customer_messages_map(all_date_folders)
    escalated_conversations = build_escalated_conversations(outlook_namespace)
    followup_sent_map = build_followup_sent_map(outlook_namespace)

    hod_popup_sent_map = build_hod_popup_sent_map(outlook_namespace)

    if INTERNAL_MAIL_DOMAINS:
        internal_rows = check_and_process_internal_mails(folder, sent_map=sent_map)
        #No excel Sheet for internal_emails
        # internal_output_file = os.path.join(
        #     PROJECT_DIRECTORY,
        #     f"internal_emails_{target_date_str}.xlsx"
        # )
        # write_to_excel(internal_rows, internal_output_file)


    
    existing_draft_conversations = (
        build_existing_draft_conversations(outlook_namespace) if IS_CUSTOMER_CARE else set()
    )

    print(f"Reading emails from '{folder.Name}' for Date : {target_date}...")
    rows = extract_conversations(
        folder, date_folder_cache, all_date_folders, sent_map, customer_map,
        seen_entry_ids, escalated_conversations, existing_draft_conversations,
        limit=LIMIT
    )
    write_to_excel(rows, OUTPUT_FILE)

    # WORK FOR LAST N DAYS REPORT
    weekly_rows = extract_conversations_last_n_days(folder, days=NO_OF_DAYS, sent_map=sent_map)
    weekly_output_file = os.path.join(
        PROJECT_DIRECTORY,
        f"emails_export_last{NO_OF_DAYS-DAYS_OFFSET}days_{target_date_str}.xlsx"
    )
    write_to_excel(weekly_rows, weekly_output_file)

    if ENABLE_FOLDER_CLEANUP:
        delete_old_date_folders(folder, days=DELETE_OLDER_THAN_DAYS)

    if IS_CUSTOMER_CARE:
        check_and_send_followup_nudges(folder, customer_map, sent_map, followup_sent_map)
        check_and_notify_hod_overdue(folder, customer_map, sent_map, hod_popup_sent_map)

    